import requests
import openpyxl
from datetime import datetime

# Nome do arquivo Excel
arquivo_excel = "cotacao_btc.xlsx"

# Tentar abrir o arquivo existente
try:
    wb = openpyxl.load_workbook(arquivo_excel)
    ws = wb.active
except FileNotFoundError:
    print("Arquivo não encontrado!")
    exit()

# Percorrer todas as moedas listadas na coluna A (exceto cabeçalho)
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
    currency = row[0].value  # Ex: BTC/USDT
    if currency:
        # Converter para formato da Binance (ex: BTCUSDT)
        symbol = currency.replace("/", "")

        # Buscar o preço na Binance
        url = f"https://api.binance.com/api/v3/ticker/price?symbol={symbol}"
        response = requests.get(url)

        if response.status_code == 200:
            data = response.json()
            preco = float(data["price"])

            # Atualizar preço (coluna B) e data da última atualização (coluna C)
            ws.cell(row=row[0].row, column=2, value=preco)
            ws.cell(row=row[0].row, column=3, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

            print(f"Atualizado {currency}: {preco}")
        else:
            print(f"Erro ao buscar {currency}")

# Salvar o arquivo atualizado
wb.save(arquivo_excel)
print("Planilha atualizada com sucesso!")
